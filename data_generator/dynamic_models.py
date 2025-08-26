import os
import json
from django.db import models, connection
from django.conf import settings
from django.core.management import call_command
from django.core.management.base import CommandError
from django.apps import apps
import importlib
from faker import Faker
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import random
import string

fake = Faker()

# Import AI data service
try:
    from .ai_data_service import get_ai_generator
    AI_AVAILABLE = True
except ImportError:
    AI_AVAILABLE = False


class DynamicModelGenerator:
    """Handle dynamic Django model creation, migration, and data generation"""
    
    FIELD_TYPE_MAPPING = {
        'string': models.CharField,
        'text': models.TextField, 
        'number': models.IntegerField,
        'decimal': models.DecimalField,
        'boolean': models.BooleanField,
        'date': models.DateField,
        'datetime': models.DateTimeField,
        'email': models.EmailField,
        'url': models.URLField,
        'list': models.TextField,  # Store as JSON string
        'choice': models.CharField,
    }
    
    def __init__(self):
        self.app_name = 'data_generator'
    
    def create_model_class(self, table_definition):
        """Create a Django model class dynamically"""
        table_name = table_definition['table_name']
        fields_definition = table_definition['fields_definition']
        
        # Create model attributes dictionary
        attrs = {
            '__module__': f'{self.app_name}.models',
            '__tablename__': table_name,
            'Meta': type('Meta', (), {
                'db_table': table_name,
                'verbose_name': table_definition.get('display_name', table_name),
                'verbose_name_plural': table_definition.get('display_name', table_name) + 's'
            }),
        }
        
        # Add fields based on field definitions
        for field_def in fields_definition:
            field_name = field_def['name']
            field_type = field_def['type']
            field_options = field_def.get('options', {})
            
            # Get Django field class
            django_field_class = self.FIELD_TYPE_MAPPING.get(field_type, models.CharField)
            
            # Create field kwargs
            field_kwargs = self._get_field_kwargs(field_type, field_options)
            
            # Add field to model
            attrs[field_name] = django_field_class(**field_kwargs)
        
        # Add primary key if not specified
        if 'id' not in [f['name'] for f in fields_definition]:
            attrs['id'] = models.AutoField(primary_key=True)
            
        # Add created_at timestamp
        attrs['created_at'] = models.DateTimeField(auto_now_add=True)
        
        # Create the model class
        model_class = type(table_name.title() + 'Model', (models.Model,), attrs)
        
        return model_class
    
    def _get_field_kwargs(self, field_type, options):
        """Get field kwargs based on field type and options"""
        kwargs = {
            'blank': options.get('nullable', False),
            'null': options.get('nullable', False),
        }
        
        if field_type == 'string':
            kwargs['max_length'] = options.get('max_length', 255)
            
        elif field_type == 'text':
            kwargs.pop('max_length', None)  # TextField doesn't need max_length
            
        elif field_type == 'number':
            if 'min_value' in options:
                # Note: Django IntegerField doesn't have min_value, we'll handle this in validation
                pass
                
        elif field_type == 'decimal':
            kwargs['max_digits'] = options.get('max_digits', 10)
            kwargs['decimal_places'] = options.get('decimal_places', 2)
            
        elif field_type == 'choice':
            choices = options.get('choices', [])
            if choices:
                kwargs['choices'] = [(choice, choice) for choice in choices]
                kwargs['max_length'] = max(len(str(choice)) for choice in choices) if choices else 50
            else:
                kwargs['max_length'] = 50
                
        elif field_type == 'list':
            # Store as text field (will be JSON)
            kwargs.pop('max_length', None)
            
        if options.get('default') is not None:
            kwargs['default'] = options['default']
            
        return kwargs
    
    def create_migration_file(self, table_definition, model_class):
        """Create a migration file for the dynamic model"""
        from django.db import connection
        from django.core.management import call_command
        
        migration_name = f"create_{table_definition['table_name']}"
        
        # Create migrations directory if it doesn't exist
        migrations_dir = os.path.join(settings.BASE_DIR, self.app_name, 'migrations')
        os.makedirs(migrations_dir, exist_ok=True)
        
        # Get the latest migration to set as dependency
        existing_migrations = [f for f in os.listdir(migrations_dir) 
                             if f.endswith('.py') and f[0].isdigit()]
        existing_migrations.sort()
        
        if existing_migrations:
            latest_migration = existing_migrations[-1].split('_')[0]
            dependency = f"('{self.app_name}', '{latest_migration}_{existing_migrations[-1].split('_', 1)[1][:-3]}')"
        else:
            dependency = f"('{self.app_name}', '0001_initial')"
        
        # Find next migration number
        next_num = len(existing_migrations) + 1
        migration_filename = f"{next_num:04d}_{migration_name}.py"
        migration_path = os.path.join(migrations_dir, migration_filename)
        
        migration_content = self._generate_migration_content_improved(
            migration_name, table_definition, model_class, dependency
        )
        
        # Write migration file
        with open(migration_path, 'w') as f:
            f.write(migration_content)
            
        return migration_filename
    
    def _generate_migration_content(self, migration_name, table_definition, model_class):
        """Generate migration file content (legacy)"""
        return f'''# Generated migration for dynamic table: {table_definition['table_name']}
from django.db import migrations, models


class Migration(migrations.Migration):

    dependencies = [
        ('{self.app_name}', '0001_initial'),
    ]

    operations = [
        migrations.RunSQL(
            """
            CREATE TABLE IF NOT EXISTS {table_definition['table_name']} (
                {self._generate_sql_fields(table_definition['fields_definition'])}
            )
            """,
            reverse_sql="DROP TABLE IF EXISTS {table_definition['table_name']}"
        ),
    ]
'''

    def _generate_migration_content_improved(self, migration_name, table_definition, model_class, dependency):
        """Generate improved migration file content with proper dependencies"""
        return f'''# Generated migration for dynamic table: {table_definition['table_name']}
from django.db import migrations, models


class Migration(migrations.Migration):

    dependencies = [
        {dependency},
    ]

    operations = [
        migrations.RunSQL(
            """
            CREATE TABLE IF NOT EXISTS {table_definition['table_name']} (
                {self._generate_sql_fields(table_definition['fields_definition'])}
            )
            """,
            reverse_sql="DROP TABLE IF EXISTS {table_definition['table_name']}"
        ),
    ]
'''
    
    def _generate_sql_fields(self, fields_definition):
        """Generate SQL field definitions"""
        sql_fields = ['id INTEGER PRIMARY KEY AUTOINCREMENT']
        
        for field_def in fields_definition:
            field_name = field_def['name']
            field_type = field_def['type']
            options = field_def.get('options', {})
            
            sql_type = self._get_sql_type(field_type, options)
            nullable = 'NULL' if options.get('nullable', False) else 'NOT NULL'
            
            sql_fields.append(f'{field_name} {sql_type} {nullable}')
        
        sql_fields.append('created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP')
        
        return ',\n                '.join(sql_fields)
    
    def _get_sql_type(self, field_type, options):
        """Get SQL type for field"""
        sql_types = {
            'string': f"VARCHAR({options.get('max_length', 255)})",
            'text': 'TEXT',
            'number': 'INTEGER',
            'decimal': f"DECIMAL({options.get('max_digits', 10)}, {options.get('decimal_places', 2)})",
            'boolean': 'BOOLEAN',
            'date': 'DATE',
            'datetime': 'DATETIME',
            'email': f"VARCHAR({options.get('max_length', 255)})",
            'url': f"VARCHAR({options.get('max_length', 500)})",
            'list': 'TEXT',
            'choice': f"VARCHAR({options.get('max_length', 50)})",
        }
        return sql_types.get(field_type, 'VARCHAR(255)')
    
    def run_migration(self, migration_filename):
        """Run the migration"""
        try:
            call_command('migrate', verbosity=0)
            return True
        except CommandError as e:
            print(f"Migration error: {e}")
            return False
    
    def run_migration_with_feedback(self, migration_filename):
        """Run the migration with detailed error feedback"""
        try:
            call_command('migrate', verbosity=1)
            return True, None
        except CommandError as e:
            error_msg = str(e)
            print(f"Migration error: {error_msg}")
            
            # Check for common migration issues
            if "Conflicting migrations detected" in error_msg:
                return False, "Migration conflict detected. Please try again or contact administrator."
            elif "relation already exists" in error_msg:
                return False, "Table already exists in database. Please choose a different table name."
            elif "syntax error" in error_msg.lower():
                return False, "SQL syntax error. Please check your field definitions."
            else:
                return False, f"Migration failed: {error_msg}"
        except Exception as e:
            error_msg = str(e)
            print(f"Unexpected migration error: {error_msg}")
            return False, f"Unexpected error: {error_msg}"
    
    def generate_synthetic_data(self, table_definition, num_records=100, openai_api_key=None):
        """Generate synthetic data for the dynamic table"""
        table_name = table_definition['table_name']
        fields_definition = table_definition['fields_definition']
        
        # Check if AI should be used
        use_ai = AI_AVAILABLE and openai_api_key and any(
            field_def.get('options', {}).get('ai_description') 
            for field_def in fields_definition
        )
        
        ai_generator = None
        if use_ai:
            try:
                ai_generator = get_ai_generator(openai_api_key)
            except Exception as e:
                print(f"Failed to initialize AI generator: {e}")
                use_ai = False
        
        data = []
        for _ in range(num_records):
            record = {}
            for field_def in fields_definition:
                field_name = field_def['name']
                field_type = field_def['type']
                options = field_def.get('options', {})
                ai_description = options.get('ai_description')
                faker_type = options.get('faker_type')  # Keep for backward compatibility
                
                # Use AI generation if available and description provided
                if use_ai and ai_description and ai_description.strip():
                    try:
                        record[field_name] = ai_generator.generate_field_value(
                            field_name, field_type, ai_description
                        )
                    except Exception as e:
                        print(f"AI generation failed for {field_name}: {e}, falling back to traditional method")
                        record[field_name] = self._generate_field_value(
                            field_type, field_name, options, faker_type
                        )
                else:
                    # Use traditional generation
                    record[field_name] = self._generate_field_value(
                        field_type, field_name, options, faker_type
                    )
            data.append(record)
        
        return data
    
    def _generate_field_value(self, field_type, field_name, options, faker_type=None):
        """Generate a single field value"""
        # Use specific faker if provided
        if faker_type:
            return self._get_faker_value(faker_type, options)
        
        # Generate based on field name heuristics
        field_name_lower = field_name.lower()
        
        if 'name' in field_name_lower:
            return fake.name()
        elif 'email' in field_name_lower:
            return fake.email()
        elif 'phone' in field_name_lower:
            return fake.phone_number()
        elif 'address' in field_name_lower:
            return fake.address()
        elif 'city' in field_name_lower:
            return fake.city()
        elif 'country' in field_name_lower:
            return fake.country()
        elif 'company' in field_name_lower:
            return fake.company()
        
        # Generate based on field type
        if field_type == 'string':
            return fake.text(max_nb_chars=options.get('max_length', 50))
        elif field_type == 'text':
            return fake.paragraph(nb_sentences=random.randint(2, 5))
        elif field_type == 'number':
            min_val = options.get('min_value', 1)
            max_val = options.get('max_value', 1000)
            return random.randint(min_val, max_val)
        elif field_type == 'decimal':
            return round(random.uniform(0, 1000), options.get('decimal_places', 2))
        elif field_type == 'boolean':
            return random.choice([True, False])
        elif field_type == 'date':
            return fake.date()
        elif field_type == 'datetime':
            return fake.date_time()
        elif field_type == 'email':
            return fake.email()
        elif field_type == 'url':
            return fake.url()
        elif field_type == 'choice':
            choices = options.get('choices', ['Option A', 'Option B', 'Option C'])
            return random.choice(choices)
        elif field_type == 'list':
            # Generate a list as JSON string
            list_items = [fake.word() for _ in range(random.randint(1, 5))]
            return json.dumps(list_items)
        else:
            return fake.word()
    
    def _get_faker_value(self, faker_type, options):
        """Get value from faker based on type"""
        faker_methods = {
            'name': fake.name,
            'first_name': fake.first_name,
            'last_name': fake.last_name,
            'email': fake.email,
            'phone': fake.phone_number,
            'address': fake.address,
            'city': fake.city,
            'country': fake.country,
            'company': fake.company,
            'job': fake.job,
            'sentence': fake.sentence,
            'paragraph': fake.paragraph,
            'uuid': fake.uuid4,
            'credit_card': fake.credit_card_number,
            'ssn': fake.ssn,
            'color': fake.color_name,
        }
        
        if faker_type in faker_methods:
            return faker_methods[faker_type]()
        return fake.word()
    
    def create_excel_file(self, table_definition, data, output_path):
        """Create Excel file with synthetic data"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = table_definition['display_name']
        
        # Create header row
        fields = table_definition['fields_definition']
        headers = [field['name'] for field in fields]
        
        # Style for headers
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data rows
        for row, record in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                value = record.get(header, '')
                if isinstance(value, list):
                    value = ', '.join(map(str, value))
                sheet.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        workbook.save(output_path)
        return output_path
    
    def insert_data_to_db(self, table_definition, data):
        """Insert synthetic data directly into the database"""
        table_name = table_definition['table_name']
        
        if not data:
            return
        
        # Get field names
        field_names = [field['name'] for field in table_definition['fields_definition']]
        
        # Prepare SQL
        placeholders = ', '.join(['?' for _ in field_names])
        fields_str = ', '.join(field_names)
        sql = f"INSERT INTO {table_name} ({fields_str}) VALUES ({placeholders})"
        
        # Insert data
        with connection.cursor() as cursor:
            for record in data:
                values = [record.get(field, None) for field in field_names]
                cursor.execute(sql, values)
